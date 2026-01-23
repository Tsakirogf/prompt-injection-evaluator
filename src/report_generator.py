"""
Report Generator Service

Generates PDF and Excel reports for prompt injection evaluation results.
"""
import os
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
)
from reportlab.lib import colors


class ReportGenerator:
    """
    Service for generating evaluation reports in PDF and Excel formats.
    """
    
    def __init__(self, output_dir: str = "reports"):
        """
        Initialize the ReportGenerator.
        
        Args:
            output_dir: Directory where reports will be saved.
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def sanitize_filename(self, name: str) -> str:
        """
        Sanitize model name for use in filename.
        
        Args:
            name: Model name to sanitize.
        
        Returns:
            Sanitized filename-safe string.
        """
        # Replace slashes and special characters
        safe_name = name.replace('/', '_').replace('\\', '_')
        safe_name = ''.join(c for c in safe_name if c.isalnum() or c in ('-', '_', '.'))
        return safe_name
    
    def generate_pdf_report(
        self,
        model_name: str,
        test_results: List[Dict[str, Any]],
        metadata: Optional[Dict[str, Any]] = None,
        stats_by_category: Optional[Dict[str, Dict[str, int]]] = None,
        stats_by_severity: Optional[Dict[str, Dict[str, int]]] = None
    ) -> Path:
        """
        Generate a PDF report for a model's test results.
        
        Args:
            model_name: Name of the model.
            test_results: List of test result dictionaries.
            metadata: Optional metadata to include in report.
        
        Returns:
            Path to the generated PDF file.
        """
        safe_name = self.sanitize_filename(model_name)
        filename = self.output_dir / f"{safe_name}.pdf"
        
        doc = SimpleDocTemplate(
            str(filename),
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph(
            f"<b>Prompt Injection Evaluation Report</b><br/>{model_name}",
            styles["Title"]
        )
        story.append(title)
        story.append(Spacer(1, 1*cm))
        
        # Metadata
        if metadata:
            story.append(Paragraph("<b>Report Information</b>", styles["Heading2"]))
            story.append(Spacer(1, 0.3*cm))
            
            for key, value in metadata.items():
                story.append(Paragraph(
                    f"<b>{key}:</b> {value}",
                    styles["Normal"]
                ))
            story.append(Spacer(1, 0.5*cm))
        
        # Summary
        if test_results:
            story.append(Paragraph("<b>Test Results Summary</b>", styles["Heading2"]))
            story.append(Spacer(1, 0.3*cm))
            
            passed = sum(1 for r in test_results if r.get('passed', False))
            failed = len(test_results) - passed
            pass_rate = (passed / len(test_results) * 100) if len(test_results) > 0 else 0
            
            story.append(Paragraph(
                f"<b>Total Tests:</b> {len(test_results)}",
                styles["Normal"]
            ))
            story.append(Paragraph(
                f"<b>Passed:</b> {passed} | <b>Failed:</b> {failed} | <b>Pass Rate:</b> {pass_rate:.1f}%",
                styles["Normal"]
            ))
            story.append(Spacer(1, 0.5*cm))
            
            # Category breakdown
            if stats_by_category:
                story.append(Paragraph("<b>Results by Attack Category:</b>", styles["Heading3"]))
                story.append(Spacer(1, 0.2*cm))
                
                for category, stats in sorted(stats_by_category.items()):
                    cat_pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
                    story.append(Paragraph(
                        f"  • {category}: {stats['passed']}/{stats['total']} passed ({cat_pass_rate:.1f}%)",
                        styles["Normal"]
                    ))
                story.append(Spacer(1, 0.3*cm))
            
            # Severity breakdown
            if stats_by_severity:
                story.append(Paragraph("<b>Results by Severity Level:</b>", styles["Heading3"]))
                story.append(Spacer(1, 0.2*cm))
                
                severity_order = ['critical', 'high', 'medium', 'low']
                for severity in severity_order:
                    if severity in stats_by_severity:
                        stats = stats_by_severity[severity]
                        sev_pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
                        story.append(Paragraph(
                            f"  • {severity}: {stats['passed']}/{stats['total']} passed ({sev_pass_rate:.1f}%)",
                            styles["Normal"]
                        ))
                story.append(Spacer(1, 0.5*cm))
        
        # Detailed Results
        story.append(Paragraph("<b>Detailed Test Results</b>", styles["Heading2"]))
        story.append(Spacer(1, 0.3*cm))
        
        for idx, result in enumerate(test_results, 1):
            story.append(Paragraph(
                f"<b>Test {idx}: {result.get('test_name', 'Unknown')}</b>",
                styles["Heading3"]
            ))
            
            story.append(Paragraph(
                f"<b>Test ID:</b> {result.get('test_id', 'N/A')}",
                styles["Normal"]
            ))
            
            story.append(Paragraph(
                f"<b>Category:</b> {result.get('category', 'N/A')}",
                styles["Normal"]
            ))
            
            story.append(Paragraph(
                f"<b>Severity:</b> {result.get('severity', 'N/A')}",
                styles["Normal"]
            ))
            
            status = "✓ PASSED" if result.get('passed', False) else "✗ FAILED"
            story.append(Paragraph(
                f"<b>Status:</b> {status}",
                styles["Normal"]
            ))
            
            if result.get('output'):
                story.append(Spacer(1, 0.2*cm))
                story.append(Paragraph("<b>Model Output:</b>", styles["Normal"]))
                output_text = str(result['output']).replace('<', '&lt;').replace('>', '&gt;')
                story.append(Paragraph(output_text[:500], styles["Code"]))
            
            if result.get('notes'):
                story.append(Spacer(1, 0.2*cm))
                story.append(Paragraph(f"<b>Notes:</b> {result['notes']}", styles["Normal"]))
            
            story.append(Spacer(1, 0.5*cm))
            
            if idx < len(test_results):
                story.append(PageBreak())
        
        # Add transcript section
        story.append(PageBreak())
        story.append(Paragraph("<b>Full Test Transcript</b>", styles["Heading2"]))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            "This section contains the complete conversation transcript for each test, "
            "including system prompts, user prompts, model responses, and expected behavior.",
            styles["Normal"]
        ))
        story.append(Spacer(1, 0.5*cm))
        
        for idx, result in enumerate(test_results, 1):
            story.append(Paragraph(
                f"<b>Test {idx}: {result.get('test_name', 'Unknown')}</b> ({result.get('test_id', 'N/A')})",
                styles["Heading3"]
            ))
            story.append(Spacer(1, 0.2*cm))
            
            # System Prompt
            story.append(Paragraph("<b>System Prompt:</b>", styles["Normal"]))
            system_prompt = str(result.get('system_prompt', 'N/A')).replace('<', '&lt;').replace('>', '&gt;')
            story.append(Paragraph(system_prompt, styles["Code"]))
            story.append(Spacer(1, 0.3*cm))
            
            # User Prompt
            story.append(Paragraph("<b>User Prompt:</b>", styles["Normal"]))
            user_prompt = str(result.get('user_prompt', 'N/A')).replace('<', '&lt;').replace('>', '&gt;')
            story.append(Paragraph(user_prompt, styles["Code"]))
            story.append(Spacer(1, 0.3*cm))
            
            # Model Response
            story.append(Paragraph("<b>Model Response:</b>", styles["Normal"]))
            output_text = str(result.get('output', 'No output')).replace('<', '&lt;').replace('>', '&gt;')
            story.append(Paragraph(output_text, styles["Code"]))
            story.append(Spacer(1, 0.3*cm))
            
            # Expected Behavior
            story.append(Paragraph("<b>Expected Behavior:</b>", styles["Normal"]))
            expected_behavior = str(result.get('expected_behavior', 'N/A')).replace('_', ' ').title()
            story.append(Paragraph(expected_behavior, styles["Normal"]))
            story.append(Spacer(1, 0.3*cm))
            
            # Status
            status = "✓ PASSED" if result.get('passed', False) else "✗ FAILED"
            status_color = 'green' if result.get('passed', False) else 'red'
            story.append(Paragraph(
                f"<b>Result:</b> <font color='{status_color}'>{status}</font>",
                styles["Normal"]
            ))
            
            if idx < len(test_results):
                story.append(Spacer(1, 1*cm))
        
        doc.build(story)
        return filename
    
    def generate_excel_report(
        self,
        model_name: str,
        test_results: List[Dict[str, Any]],
        metadata: Optional[Dict[str, Any]] = None,
        stats_by_category: Optional[Dict[str, Dict[str, int]]] = None,
        stats_by_severity: Optional[Dict[str, Dict[str, int]]] = None
    ) -> Path:
        """
        Generate an Excel report for a model's test results.
        
        Args:
            model_name: Name of the model.
            test_results: List of test result dictionaries.
            metadata: Optional metadata to include in report.
        
        Returns:
            Path to the generated Excel file.
        """
        safe_name = self.sanitize_filename(model_name)
        filename = self.output_dir / f"{safe_name}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Evaluation Results"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Metadata section
        row = 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = f"Prompt Injection Evaluation Report: {model_name}"
        ws[f'A{row}'].font = Font(size=16, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        if metadata:
            for key, value in metadata.items():
                ws[f'A{row}'] = key
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
            row += 1
        
        # Summary section
        if test_results:
            passed = sum(1 for r in test_results if r.get('passed', False))
            failed = len(test_results) - passed
            pass_rate = (passed / len(test_results) * 100) if len(test_results) > 0 else 0
            
            ws[f'A{row}'] = "Total Tests"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = len(test_results)
            row += 1
            
            ws[f'A{row}'] = "Passed"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = passed
            ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            row += 1
            
            ws[f'A{row}'] = "Failed"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = failed
            ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            row += 1
            
            ws[f'A{row}'] = "Pass Rate"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = f"{pass_rate:.1f}%"
            row += 2
            
            # Category breakdown
            if stats_by_category:
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Results by Attack Category"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                for category, stats in sorted(stats_by_category.items()):
                    cat_pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
                    ws[f'A{row}'] = category
                    ws[f'B{row}'] = f"{stats['passed']}/{stats['total']} ({cat_pass_rate:.1f}%)"
                    row += 1
                row += 1
            
            # Severity breakdown
            if stats_by_severity:
                ws.merge_cells(f'A{row}:B{row}')
                ws[f'A{row}'] = "Results by Severity Level"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                severity_order = ['critical', 'high', 'medium', 'low']
                for severity in severity_order:
                    if severity in stats_by_severity:
                        stats = stats_by_severity[severity]
                        sev_pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
                        ws[f'A{row}'] = severity
                        ws[f'B{row}'] = f"{stats['passed']}/{stats['total']} ({sev_pass_rate:.1f}%)"
                        row += 1
                row += 1
        
        # Results table headers
        headers = ['Test ID', 'Test Name', 'Category', 'Severity', 'Status', 'Output', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Results data
        for result in test_results:
            ws.cell(row=row, column=1, value=result.get('test_id', 'N/A'))
            ws.cell(row=row, column=2, value=result.get('test_name', 'Unknown'))
            ws.cell(row=row, column=3, value=result.get('category', 'N/A'))
            ws.cell(row=row, column=4, value=result.get('severity', 'N/A'))
            
            status = "PASSED" if result.get('passed', False) else "FAILED"
            status_cell = ws.cell(row=row, column=5, value=status)
            if status == "PASSED":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws.cell(row=row, column=6, value=str(result.get('output', ''))[:500])
            ws.cell(row=row, column=7, value=result.get('notes', ''))
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 30
        
        # Add transcript sheet
        ws_transcript = wb.create_sheet("Transcript")
        row = 1
        
        # Title
        ws_transcript.merge_cells(f'A{row}:D{row}')
        ws_transcript[f'A{row}'] = "Full Test Transcript"
        ws_transcript[f'A{row}'].font = Font(size=16, bold=True)
        ws_transcript[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        ws_transcript[f'A{row}'] = "This sheet contains the complete conversation transcript for each test."
        ws_transcript[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 2
        
        for idx, result in enumerate(test_results, 1):
            # Test header
            ws_transcript.merge_cells(f'A{row}:D{row}')
            ws_transcript[f'A{row}'] = f"Test {idx}: {result.get('test_name', 'Unknown')} ({result.get('test_id', 'N/A')})"
            ws_transcript[f'A{row}'].font = Font(bold=True, size=12)
            ws_transcript[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            row += 1
            
            # System Prompt
            ws_transcript[f'A{row}'] = "System Prompt:"
            ws_transcript[f'A{row}'].font = Font(bold=True)
            ws_transcript[f'A{row}'].alignment = Alignment(vertical='top')
            ws_transcript.merge_cells(f'B{row}:D{row}')
            ws_transcript[f'B{row}'] = result.get('system_prompt', 'N/A')
            ws_transcript[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
            
            # User Prompt
            ws_transcript[f'A{row}'] = "User Prompt:"
            ws_transcript[f'A{row}'].font = Font(bold=True)
            ws_transcript[f'A{row}'].alignment = Alignment(vertical='top')
            ws_transcript.merge_cells(f'B{row}:D{row}')
            ws_transcript[f'B{row}'] = result.get('user_prompt', 'N/A')
            ws_transcript[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
            
            # Model Response
            ws_transcript[f'A{row}'] = "Model Response:"
            ws_transcript[f'A{row}'].font = Font(bold=True)
            ws_transcript[f'A{row}'].alignment = Alignment(vertical='top')
            ws_transcript.merge_cells(f'B{row}:D{row}')
            ws_transcript[f'B{row}'] = result.get('output', 'No output')
            ws_transcript[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
            
            # Expected Behavior
            ws_transcript[f'A{row}'] = "Expected Behavior:"
            ws_transcript[f'A{row}'].font = Font(bold=True)
            ws_transcript[f'A{row}'].alignment = Alignment(vertical='top')
            ws_transcript.merge_cells(f'B{row}:D{row}')
            expected_behavior = str(result.get('expected_behavior', 'N/A')).replace('_', ' ').title()
            ws_transcript[f'B{row}'] = expected_behavior
            ws_transcript[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
            
            # Result
            ws_transcript[f'A{row}'] = "Result:"
            ws_transcript[f'A{row}'].font = Font(bold=True)
            status = "PASSED" if result.get('passed', False) else "FAILED"
            ws_transcript[f'B{row}'] = status
            ws_transcript[f'B{row}'].font = Font(bold=True)
            if status == "PASSED":
                ws_transcript[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws_transcript[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            row += 2
        
        # Adjust column widths for transcript sheet
        ws_transcript.column_dimensions['A'].width = 20
        ws_transcript.column_dimensions['B'].width = 80
        ws_transcript.column_dimensions['C'].width = 15
        ws_transcript.column_dimensions['D'].width = 15
        
        wb.save(filename)
        return filename
    
    def generate_reports(
        self,
        model_name: str,
        test_results: List[Dict[str, Any]],
        metadata: Optional[Dict[str, Any]] = None,
        stats_by_category: Optional[Dict[str, Dict[str, int]]] = None,
        stats_by_severity: Optional[Dict[str, Dict[str, int]]] = None
    ) -> Dict[str, Path]:
        """
        Generate both PDF and Excel reports.
        
        Args:
            model_name: Name of the model.
            test_results: List of test result dictionaries.
            metadata: Optional metadata to include in reports.
            stats_by_category: Statistics broken down by attack category.
            stats_by_severity: Statistics broken down by severity level.
        
        Returns:
            Dictionary with 'pdf' and 'excel' keys containing file paths.
        """
        pdf_path = self.generate_pdf_report(
            model_name, test_results, metadata, stats_by_category, stats_by_severity
        )
        excel_path = self.generate_excel_report(
            model_name, test_results, metadata, stats_by_category, stats_by_severity
        )
        
        return {
            'pdf': pdf_path,
            'excel': excel_path
        }
    
    def generate_comparison_report(
        self,
        all_model_results: List[Dict[str, Any]],
        test_suite: Any,
        timestamp: str
    ) -> Dict[str, Path]:
        """
        Generate comparison reports showing all models side-by-side.
        
        Args:
            all_model_results: List of result dictionaries for all models
            test_suite: TestSuite object with test metadata
            timestamp: Timestamp string for the report
        
        Returns:
            Dictionary with 'pdf' and 'excel' keys containing file paths.
        """
        pdf_path = self._generate_comparison_pdf(all_model_results, test_suite, timestamp)
        excel_path = self._generate_comparison_excel(all_model_results, test_suite, timestamp)
        
        return {
            'pdf': pdf_path,
            'excel': excel_path
        }
    
    def _generate_comparison_pdf(
        self,
        all_model_results: List[Dict[str, Any]],
        test_suite: Any,
        timestamp: str
    ) -> Path:
        """Generate PDF comparison report."""
        filename = self.output_dir / "model_comparison.pdf"
        
        doc = SimpleDocTemplate(
            str(filename),
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph(
            f"<b>Model Comparison Report</b><br/>Prompt Injection Evaluation",
            styles["Title"]
        )
        story.append(title)
        story.append(Spacer(1, 0.5*cm))
        
        # Metadata
        story.append(Paragraph(f"<b>Date:</b> {timestamp}", styles["Normal"]))
        story.append(Paragraph(f"<b>Test Suite:</b> {test_suite.name} v{test_suite.version}", styles["Normal"]))
        story.append(Paragraph(f"<b>Models Evaluated:</b> {len(all_model_results)}", styles["Normal"]))
        story.append(Spacer(1, 1*cm))
        
        # Overall comparison table
        story.append(Paragraph("<b>Overall Performance Comparison</b>", styles["Heading2"]))
        story.append(Spacer(1, 0.3*cm))
        
        comparison_data = [['Model', 'Total Tests', 'Passed', 'Failed', 'Pass Rate']]
        for result in all_model_results:
            comparison_data.append([
                result['model_name'],
                str(result['total_tests']),
                str(result['passed_tests']),
                str(result['failed_tests']),
                f"{result['pass_rate']:.1f}%"
            ])
        
        comparison_table = Table(comparison_data, colWidths=[6*cm, 2.5*cm, 2*cm, 2*cm, 2.5*cm])
        comparison_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(comparison_table)
        story.append(Spacer(1, 1*cm))
        
        # Category comparison
        story.append(Paragraph("<b>Performance by Attack Category</b>", styles["Heading2"]))
        story.append(Spacer(1, 0.3*cm))
        
        # Collect all categories
        all_categories = set()
        for result in all_model_results:
            all_categories.update(result['stats_by_category'].keys())
        
        for category in sorted(all_categories):
            story.append(Paragraph(f"<b>{category}</b>", styles["Heading3"]))
            category_data = [['Model', 'Passed', 'Failed', 'Pass Rate']]
            
            for result in all_model_results:
                if category in result['stats_by_category']:
                    stats = result['stats_by_category'][category]
                    pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
                    category_data.append([
                        result['model_name'],
                        f"{stats['passed']}/{stats['total']}",
                        str(stats['failed']),
                        f"{pass_rate:.1f}%"
                    ])
            
            category_table = Table(category_data, colWidths=[6*cm, 3*cm, 3*cm, 3*cm])
            category_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(category_table)
            story.append(Spacer(1, 0.5*cm))
        
        story.append(PageBreak())
        
        # Severity comparison
        story.append(Paragraph("<b>Performance by Severity Level</b>", styles["Heading2"]))
        story.append(Spacer(1, 0.3*cm))
        
        severity_order = ['critical', 'high', 'medium', 'low']
        for severity in severity_order:
            # Check if any model has this severity
            has_severity = any(
                severity in result['stats_by_severity'] 
                for result in all_model_results
            )
            
            if has_severity:
                story.append(Paragraph(f"<b>{severity.upper()}</b>", styles["Heading3"]))
                severity_data = [['Model', 'Passed', 'Failed', 'Pass Rate']]
                
                for result in all_model_results:
                    if severity in result['stats_by_severity']:
                        stats = result['stats_by_severity'][severity]
                        pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
                        severity_data.append([
                            result['model_name'],
                            f"{stats['passed']}/{stats['total']}",
                            str(stats['failed']),
                            f"{pass_rate:.1f}%"
                        ])
                
                severity_table = Table(severity_data, colWidths=[6*cm, 3*cm, 3*cm, 3*cm])
                severity_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(severity_table)
                story.append(Spacer(1, 0.5*cm))
        
        doc.build(story)
        return filename
    
    def _generate_comparison_excel(
        self,
        all_model_results: List[Dict[str, Any]],
        test_suite: Any,
        timestamp: str
    ) -> Path:
        """Generate Excel comparison report."""
        filename = self.output_dir / "model_comparison.xlsx"
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Title and metadata
        row = 1
        ws_summary.merge_cells(f'A{row}:E{row}')
        ws_summary[f'A{row}'] = "Model Comparison Report - Prompt Injection Evaluation"
        ws_summary[f'A{row}'].font = Font(size=16, bold=True)
        ws_summary[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        ws_summary[f'A{row}'] = "Date:"
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = timestamp
        row += 1
        
        ws_summary[f'A{row}'] = "Test Suite:"
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = f"{test_suite.name} v{test_suite.version}"
        row += 1
        
        ws_summary[f'A{row}'] = "Models Evaluated:"
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = len(all_model_results)
        row += 2
        
        # Overall comparison table
        headers = ['Model', 'Total Tests', 'Passed', 'Failed', 'Pass Rate']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for result in all_model_results:
            ws_summary.cell(row=row, column=1, value=result['model_name'])
            ws_summary.cell(row=row, column=2, value=result['total_tests'])
            ws_summary.cell(row=row, column=3, value=result['passed_tests'])
            ws_summary.cell(row=row, column=4, value=result['failed_tests'])
            ws_summary.cell(row=row, column=5, value=f"{result['pass_rate']:.1f}%")
            row += 1
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 12
        ws_summary.column_dimensions['D'].width = 12
        ws_summary.column_dimensions['E'].width = 15
        
        # Category comparison sheet
        ws_category = wb.create_sheet("By Category")
        row = 1
        
        ws_category.merge_cells(f'A{row}:E{row}')
        ws_category[f'A{row}'] = "Performance by Attack Category"
        ws_category[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Collect all categories
        all_categories = set()
        for result in all_model_results:
            all_categories.update(result['stats_by_category'].keys())
        
        for category in sorted(all_categories):
            ws_category.merge_cells(f'A{row}:E{row}')
            ws_category[f'A{row}'] = category
            ws_category[f'A{row}'].font = Font(bold=True, size=12)
            ws_category[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            row += 1
            
            headers = ['Model', 'Total', 'Passed', 'Failed', 'Pass Rate']
            for col, header in enumerate(headers, 1):
                cell = ws_category.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            for result in all_model_results:
                if category in result['stats_by_category']:
                    stats = result['stats_by_category'][category]
                    pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
                    ws_category.cell(row=row, column=1, value=result['model_name'])
                    ws_category.cell(row=row, column=2, value=stats['total'])
                    ws_category.cell(row=row, column=3, value=stats['passed'])
                    ws_category.cell(row=row, column=4, value=stats['failed'])
                    ws_category.cell(row=row, column=5, value=f"{pass_rate:.1f}%")
                    row += 1
            row += 1
        
        ws_category.column_dimensions['A'].width = 40
        ws_category.column_dimensions['B'].width = 12
        ws_category.column_dimensions['C'].width = 12
        ws_category.column_dimensions['D'].width = 12
        ws_category.column_dimensions['E'].width = 15
        
        # Severity comparison sheet
        ws_severity = wb.create_sheet("By Severity")
        row = 1
        
        ws_severity.merge_cells(f'A{row}:E{row}')
        ws_severity[f'A{row}'] = "Performance by Severity Level"
        ws_severity[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        severity_order = ['critical', 'high', 'medium', 'low']
        for severity in severity_order:
            # Check if any model has this severity
            has_severity = any(
                severity in result['stats_by_severity'] 
                for result in all_model_results
            )
            
            if has_severity:
                ws_severity.merge_cells(f'A{row}:E{row}')
                ws_severity[f'A{row}'] = severity.upper()
                ws_severity[f'A{row}'].font = Font(bold=True, size=12)
                ws_severity[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                row += 1
                
                headers = ['Model', 'Total', 'Passed', 'Failed', 'Pass Rate']
                for col, header in enumerate(headers, 1):
                    cell = ws_severity.cell(row=row, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                row += 1
                
                for result in all_model_results:
                    if severity in result['stats_by_severity']:
                        stats = result['stats_by_severity'][severity]
                        pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
                        ws_severity.cell(row=row, column=1, value=result['model_name'])
                        ws_severity.cell(row=row, column=2, value=stats['total'])
                        ws_severity.cell(row=row, column=3, value=stats['passed'])
                        ws_severity.cell(row=row, column=4, value=stats['failed'])
                        ws_severity.cell(row=row, column=5, value=f"{pass_rate:.1f}%")
                        row += 1
                row += 1
        
        ws_severity.column_dimensions['A'].width = 40
        ws_severity.column_dimensions['B'].width = 12
        ws_severity.column_dimensions['C'].width = 12
        ws_severity.column_dimensions['D'].width = 12
        ws_severity.column_dimensions['E'].width = 15
        
        wb.save(filename)
        return filename
