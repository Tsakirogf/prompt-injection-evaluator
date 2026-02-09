"""
Response Collector Module

Collects model responses WITHOUT evaluation, enabling:
- Re-evaluation without endpoint costs
- Testing different evaluators on same responses
- Storing raw responses for analysis

Phase 1 of the separated workflow:
TestSuite -> ResponseCollector -> responses.xlsx
"""
import json
import time
from datetime import datetime
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from test_suite_loader import TestSuite, TestCase

# Check for openpyxl availability
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class CollectedResponse:
    """
    Represents a single collected response from a model.
    Contains all data needed for later evaluation.
    """
    test_id: str
    test_name: str
    category: str
    severity: str
    system_prompt: str
    user_prompt: str
    model_response: str
    response_timestamp: str
    response_time_ms: Optional[float] = None
    error: Optional[str] = None

    # Test configuration for re-evaluation
    expected_behavior: str = ""
    protected_secret: Optional[str] = None
    expected_keywords_absent: List[str] = field(default_factory=list)
    expected_keywords_present: List[str] = field(default_factory=list)


@dataclass
class CollectionMetadata:
    """Metadata about a response collection run."""
    model_name: str
    endpoint_url: Optional[str]
    test_suite_name: str
    test_suite_version: str
    started_at: str
    completed_at: Optional[str] = None
    total_tests: int = 0
    successful_responses: int = 0
    failed_responses: int = 0


class ResponseCollector:
    """
    Collects model responses without performing evaluation.

    This enables separating the expensive endpoint calls from
    the evaluation logic, allowing re-evaluation without costs.
    """

    def __init__(self, model_config: Dict[str, Any]):
        """
        Initialize the ResponseCollector.

        Args:
            model_config: Model configuration dictionary
        """
        self.model_config = model_config
        self.model_name = model_config['name']
        self.responses: List[CollectedResponse] = []
        self.metadata: Optional[CollectionMetadata] = None

    def collect_all(
        self,
        test_suite: TestSuite,
        verbose: bool = True
    ) -> List[CollectedResponse]:
        """
        Collect responses for all test cases without evaluation.

        Args:
            test_suite: TestSuite containing test cases
            verbose: Whether to print progress

        Returns:
            List of CollectedResponse objects
        """
        # Lazy imports - only needed for collection, not for loading xlsx
        from model_inference import ModelInference
        from endpoint_manager import EndpointManager

        if verbose:
            print(f"\n Collecting responses from: {self.model_name}")
            print("-" * 70)

        # Initialize metadata
        self.metadata = CollectionMetadata(
            model_name=self.model_name,
            endpoint_url=self.model_config.get('endpoint_url'),
            test_suite_name=test_suite.name,
            test_suite_version=test_suite.version,
            started_at=datetime.now().isoformat(),
            total_tests=len(test_suite)
        )

        # Check if this is a managed endpoint
        endpoint_manager = None
        is_managed_endpoint = (
            self.model_config.get('remote_type') == 'hf_inference_endpoint' and
            self.model_config.get('endpoint_url')
        )

        if is_managed_endpoint:
            endpoint_name = self.model_config.get('endpoint_name')
            namespace = self.model_config.get('endpoint_namespace', 'tsakirogf')

            if endpoint_name:
                if verbose:
                    print(f"  Managed endpoint detected: {endpoint_name}")
                endpoint_manager = EndpointManager(endpoint_name, namespace=namespace)
            else:
                if verbose:
                    print(f"  Finding endpoint from URL...")
                endpoint_manager = EndpointManager.from_url(
                    self.model_config['endpoint_url'],
                    namespace=namespace
                )

            if endpoint_manager:
                try:
                    if not endpoint_manager.resume(wait=True, max_wait=120):
                        if verbose:
                            print(f"  Failed to start endpoint, will attempt to continue...")
                except Exception as e:
                    if verbose:
                        print(f"  Endpoint management failed: {e}")
                        print(f"  Continuing without endpoint management...")
                    endpoint_manager = None

        # Initialize model inference
        model_inference = ModelInference(self.model_config)
        try:
            model_inference.load()
        except Exception as e:
            if endpoint_manager:
                endpoint_manager.pause()
            raise RuntimeError(f"Failed to load model: {e}")

        self.responses = []

        for idx, test_case in enumerate(test_suite, 1):
            if verbose:
                print(f"  [{idx}/{len(test_suite)}] {test_case.id} - {test_case.name}", end="")

            start_time = time.time()
            error = None
            model_response = ""

            try:
                model_response = model_inference.generate(
                    test_case.system_prompt,
                    test_case.get_user_prompt()
                )
                if model_response.startswith("[ERROR]"):
                    error = model_response
                    self.metadata.failed_responses += 1
                else:
                    self.metadata.successful_responses += 1
            except Exception as e:
                error = f"[ERROR] {str(e)}"
                model_response = error
                self.metadata.failed_responses += 1

            response_time_ms = (time.time() - start_time) * 1000

            response = CollectedResponse(
                test_id=test_case.id,
                test_name=test_case.name,
                category=test_case.category,
                severity=test_case.severity,
                system_prompt=test_case.system_prompt,
                user_prompt=test_case.get_user_prompt(),
                model_response=model_response,
                response_timestamp=datetime.now().isoformat(),
                response_time_ms=response_time_ms,
                error=error,
                expected_behavior=test_case.expected_behavior,
                protected_secret=test_case.protected_secret,
                expected_keywords_absent=test_case.expected_keywords_absent or [],
                expected_keywords_present=test_case.expected_keywords_present or []
            )

            self.responses.append(response)

            if verbose:
                status = " (error)" if error else ""
                print(f" [{response_time_ms:.0f}ms]{status}")

        # Unload model
        model_inference.unload()
        if verbose:
            print(f"  Model unloaded")

        # Pause endpoint if we started it
        if endpoint_manager:
            endpoint_manager.pause()

        # Update metadata
        self.metadata.completed_at = datetime.now().isoformat()

        if verbose:
            print(f"\n  Summary: {self.metadata.successful_responses}/{self.metadata.total_tests} successful")

        return self.responses

    def save_to_xlsx(self, output_path: Path) -> Path:
        """
        Save collected responses to xlsx file.

        Args:
            output_path: Path for output xlsx file

        Returns:
            Path to the created xlsx file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required to save xlsx files. "
                "Install with: pip install openpyxl"
            )

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Sheet 1: Responses
        ws_responses = wb.active
        ws_responses.title = "Responses"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Response headers
        response_headers = [
            'test_id', 'test_name', 'category', 'severity',
            'system_prompt', 'user_prompt', 'model_response',
            'response_timestamp', 'response_time_ms', 'error'
        ]

        for col, header in enumerate(response_headers, 1):
            cell = ws_responses.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Response data
        for row_idx, response in enumerate(self.responses, 2):
            ws_responses.cell(row=row_idx, column=1, value=response.test_id)
            ws_responses.cell(row=row_idx, column=2, value=response.test_name)
            ws_responses.cell(row=row_idx, column=3, value=response.category)
            ws_responses.cell(row=row_idx, column=4, value=response.severity)
            ws_responses.cell(row=row_idx, column=5, value=response.system_prompt)
            ws_responses.cell(row=row_idx, column=6, value=response.user_prompt)
            ws_responses.cell(row=row_idx, column=7, value=response.model_response)
            ws_responses.cell(row=row_idx, column=8, value=response.response_timestamp)
            ws_responses.cell(row=row_idx, column=9, value=response.response_time_ms)
            ws_responses.cell(row=row_idx, column=10, value=response.error)

        # Adjust column widths for Responses sheet
        ws_responses.column_dimensions['A'].width = 12
        ws_responses.column_dimensions['B'].width = 30
        ws_responses.column_dimensions['C'].width = 20
        ws_responses.column_dimensions['D'].width = 12
        ws_responses.column_dimensions['E'].width = 50
        ws_responses.column_dimensions['F'].width = 50
        ws_responses.column_dimensions['G'].width = 60
        ws_responses.column_dimensions['H'].width = 25
        ws_responses.column_dimensions['I'].width = 15
        ws_responses.column_dimensions['J'].width = 40

        # Sheet 2: Metadata
        ws_metadata = wb.create_sheet("Metadata")

        metadata_headers = ['Field', 'Value']
        for col, header in enumerate(metadata_headers, 1):
            cell = ws_metadata.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        if self.metadata:
            metadata_fields = [
                ('model_name', self.metadata.model_name),
                ('endpoint_url', self.metadata.endpoint_url or ''),
                ('test_suite_name', self.metadata.test_suite_name),
                ('test_suite_version', self.metadata.test_suite_version),
                ('started_at', self.metadata.started_at),
                ('completed_at', self.metadata.completed_at or ''),
                ('total_tests', self.metadata.total_tests),
                ('successful_responses', self.metadata.successful_responses),
                ('failed_responses', self.metadata.failed_responses),
            ]

            for row_idx, (field, value) in enumerate(metadata_fields, 2):
                ws_metadata.cell(row=row_idx, column=1, value=field)
                ws_metadata.cell(row=row_idx, column=2, value=value)

        ws_metadata.column_dimensions['A'].width = 25
        ws_metadata.column_dimensions['B'].width = 60

        # Sheet 3: TestConfig (for re-evaluation)
        ws_config = wb.create_sheet("TestConfig")

        config_headers = [
            'test_id', 'expected_behavior', 'protected_secret',
            'expected_keywords_absent', 'expected_keywords_present'
        ]

        for col, header in enumerate(config_headers, 1):
            cell = ws_config.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, response in enumerate(self.responses, 2):
            ws_config.cell(row=row_idx, column=1, value=response.test_id)
            ws_config.cell(row=row_idx, column=2, value=response.expected_behavior)
            ws_config.cell(row=row_idx, column=3, value=response.protected_secret or '')
            ws_config.cell(row=row_idx, column=4, value=json.dumps(response.expected_keywords_absent))
            ws_config.cell(row=row_idx, column=5, value=json.dumps(response.expected_keywords_present))

        ws_config.column_dimensions['A'].width = 12
        ws_config.column_dimensions['B'].width = 40
        ws_config.column_dimensions['C'].width = 25
        ws_config.column_dimensions['D'].width = 40
        ws_config.column_dimensions['E'].width = 40

        wb.save(output_path)
        return output_path

    @staticmethod
    def load_from_xlsx(xlsx_path: Path) -> Tuple[CollectionMetadata, List[CollectedResponse]]:
        """
        Load collected responses from xlsx file.

        Args:
            xlsx_path: Path to xlsx file

        Returns:
            Tuple of (CollectionMetadata, List[CollectedResponse])
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required to load xlsx files. "
                "Install with: pip install openpyxl"
            )

        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            raise FileNotFoundError(f"Response file not found: {xlsx_path}")

        wb = load_workbook(xlsx_path, data_only=True)

        # Load metadata
        ws_metadata = wb["Metadata"]
        metadata_dict = {}
        for row in ws_metadata.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0]:
                metadata_dict[row[0]] = row[1]

        metadata = CollectionMetadata(
            model_name=metadata_dict.get('model_name', ''),
            endpoint_url=metadata_dict.get('endpoint_url'),
            test_suite_name=metadata_dict.get('test_suite_name', ''),
            test_suite_version=metadata_dict.get('test_suite_version', ''),
            started_at=metadata_dict.get('started_at', ''),
            completed_at=metadata_dict.get('completed_at'),
            total_tests=int(metadata_dict.get('total_tests', 0)),
            successful_responses=int(metadata_dict.get('successful_responses', 0)),
            failed_responses=int(metadata_dict.get('failed_responses', 0))
        )

        # Load test config
        ws_config = wb["TestConfig"]
        config_dict = {}
        for row in ws_config.iter_rows(min_row=2, values_only=True):
            if row[0]:
                test_id = row[0]
                config_dict[test_id] = {
                    'expected_behavior': row[1] or '',
                    'protected_secret': row[2] if row[2] else None,
                    'expected_keywords_absent': json.loads(row[3]) if row[3] else [],
                    'expected_keywords_present': json.loads(row[4]) if row[4] else []
                }

        # Load responses
        ws_responses = wb["Responses"]
        responses = []

        for row in ws_responses.iter_rows(min_row=2, values_only=True):
            if row[0]:  # test_id exists
                test_id = row[0]
                config = config_dict.get(test_id, {})

                response = CollectedResponse(
                    test_id=test_id,
                    test_name=row[1] or '',
                    category=row[2] or '',
                    severity=row[3] or '',
                    system_prompt=row[4] or '',
                    user_prompt=row[5] or '',
                    model_response=row[6] or '',
                    response_timestamp=row[7] or '',
                    response_time_ms=float(row[8]) if row[8] else None,
                    error=row[9] if row[9] else None,
                    expected_behavior=config.get('expected_behavior', ''),
                    protected_secret=config.get('protected_secret'),
                    expected_keywords_absent=config.get('expected_keywords_absent', []),
                    expected_keywords_present=config.get('expected_keywords_present', [])
                )
                responses.append(response)

        return metadata, responses
