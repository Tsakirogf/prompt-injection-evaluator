#!/usr/bin/env python3
"""
Generate Comparison Report

Creates a comparison report from existing evaluation results.
Reads Excel files from reports/ directory and generates a comparison PDF and Excel.

Usage:
    python generate_comparison.py
    python generate_comparison.py --models "model1" "model2" "model3"
"""
import argparse
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / 'src'))

from response_collector import ResponseCollector
from test_suite_loader import TestSuiteLoader
from report_generator import ReportGenerator


def load_evaluation_results(xlsx_path: Path) -> Dict[str, Any]:
    """
    Load evaluation results from an Excel report file.

    Args:
        xlsx_path: Path to the evaluation report Excel file

    Returns:
        Dictionary with model results
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Report file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)

    # Get the first sheet (Results)
    ws_results = wb.active

    # Read metadata from the top section
    model_name = None
    total_tests = 0
    passed_tests = 0
    failed_tests = 0
    pass_rate = 0.0

    # Find where the actual results table starts
    results_start_row = None
    for idx, row in enumerate(ws_results.iter_rows(min_row=1, max_row=100, values_only=True), 1):
        if row[0]:
            # Row 1 has the model name directly (not with "Model:" label)
            if idx == 1 and isinstance(row[0], str):
                model_name = row[0]
                # Clean up the model name by removing title prefixes
                if ":" in model_name:
                    model_name = model_name.split(":", 1)[1].strip()
            # Look for metadata labels
            elif isinstance(row[0], str):
                if row[0] == "Total Tests":
                    total_tests = int(row[1]) if row[1] else 0
                elif row[0] == "Passed":
                    passed_tests = int(row[1]) if row[1] else 0
                elif row[0] == "Failed":
                    failed_tests = int(row[1]) if row[1] else 0
                elif row[0] == "Pass Rate":
                    pass_rate_str = str(row[1]).replace('%', '').strip() if row[1] else '0'
                    try:
                        pass_rate = float(pass_rate_str)
                    except:
                        pass_rate = 0.0
                # Look for table header
                elif row[0] == "Test ID":
                    results_start_row = idx + 1
                    break

    # Load test results
    test_results = []
    if results_start_row:
        for row in ws_results.iter_rows(min_row=results_start_row, values_only=True):
            if row[0]:  # test_id exists
                # Columns: Test ID, Test Name, Category, Severity, Status, Output, Notes
                result = {
                    'test_id': row[0],
                    'test_name': row[1] if len(row) > 1 else '',
                    'category': row[2] if len(row) > 2 else '',
                    'severity': row[3] if len(row) > 3 else '',
                    'passed': str(row[4]).upper() in ['PASS', 'PASSED', 'TRUE'] if len(row) > 4 else False,
                    'output': str(row[5]) if len(row) > 5 else '',
                    'notes': str(row[6]) if len(row) > 6 else ''
                }
                test_results.append(result)

    # Calculate stats by category and severity
    stats_by_category = {}
    stats_by_severity = {}

    for result in test_results:
        category = result['category']
        severity = result['severity']
        passed = result['passed']

        # Category stats
        if category not in stats_by_category:
            stats_by_category[category] = {'passed': 0, 'failed': 0, 'total': 0}
        stats_by_category[category]['total'] += 1
        if passed:
            stats_by_category[category]['passed'] += 1
        else:
            stats_by_category[category]['failed'] += 1

        # Severity stats
        if severity not in stats_by_severity:
            stats_by_severity[severity] = {'passed': 0, 'failed': 0, 'total': 0}
        stats_by_severity[severity]['total'] += 1
        if passed:
            stats_by_severity[severity]['passed'] += 1
        else:
            stats_by_severity[severity]['failed'] += 1

    return {
        'model_name': model_name,
        'total_tests': total_tests,
        'passed_tests': passed_tests,
        'failed_tests': failed_tests,
        'pass_rate': pass_rate,
        'test_results': test_results,
        'stats_by_category': stats_by_category,
        'stats_by_severity': stats_by_severity
    }


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Generate comparison report from existing evaluation results",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate comparison for all models in reports/
  python generate_comparison.py

  # Generate comparison for specific models
  python generate_comparison.py --models "meta-llama/Llama-3.1-8B-Instruct" "meta-llama/Meta-Llama-3-70B-Instruct"
        """
    )

    parser.add_argument(
        '--models', '-m',
        nargs='+',
        help='List of model names to compare (uses report filenames)'
    )

    parser.add_argument(
        '--output', '-o',
        default='reports',
        help='Output directory for comparison reports (default: reports/)'
    )

    args = parser.parse_args()

    print("\n" + "=" * 70)
    print("GENERATING MODEL COMPARISON REPORT")
    print("=" * 70)

    reports_dir = Path('reports')
    if not reports_dir.exists():
        print(f"\nError: Reports directory not found: {reports_dir}")
        sys.exit(1)

    # Find all Excel report files
    if args.models:
        # Use specified models
        excel_files = []
        for model_name in args.models:
            safe_name = model_name.replace('/', '_')
            xlsx_path = reports_dir / f"{safe_name}.xlsx"
            if xlsx_path.exists():
                excel_files.append(xlsx_path)
            else:
                print(f"Warning: Report not found for model: {model_name}")
    else:
        # Use all Excel files in reports directory
        excel_files = list(reports_dir.glob('*.xlsx'))
        # Exclude files that look like responses or comparison reports
        excel_files = [f for f in excel_files if '_responses' not in f.name and 'comparison' not in f.name.lower()]

    if not excel_files:
        print("\nError: No evaluation report Excel files found in reports/")
        print("Please run evaluations first using: python src/main.py --model <model_name>")
        sys.exit(1)

    print(f"\nFound {len(excel_files)} model evaluation reports:")
    for xlsx_file in excel_files:
        print(f"  - {xlsx_file.stem}")

    print("\nLoading evaluation results...")
    all_model_results = []

    for xlsx_file in excel_files:
        try:
            print(f"  Loading {xlsx_file.name}...", end=" ")
            result = load_evaluation_results(xlsx_file)
            all_model_results.append(result)
            print(f"OK ({result['passed_tests']}/{result['total_tests']} passed)")
        except Exception as e:
            print(f"FAILED: {e}")

    if not all_model_results:
        print("\nError: Failed to load any evaluation results")
        sys.exit(1)

    # Load test suite for metadata
    print("\nLoading test suite metadata...")
    test_suite_loader = TestSuiteLoader('config/prompt_cases')
    test_suite = test_suite_loader.load()

    # Generate comparison report
    print("\n" + "=" * 70)
    print("GENERATING COMPARISON REPORT")
    print("=" * 70)

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    report_generator = ReportGenerator(str(output_dir))
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    comparison_report = report_generator.generate_comparison_report(
        all_model_results,
        test_suite,
        timestamp
    )

    print(f"\n[OK] PDF:   {comparison_report['pdf']}")
    print(f"[OK] Excel: {comparison_report['excel']}")

    print("\n" + "=" * 70)
    print("COMPARISON SUMMARY")
    print("=" * 70)

    # Sort by pass rate (descending)
    sorted_results = sorted(all_model_results, key=lambda x: x['pass_rate'], reverse=True)

    print(f"\n{'Model':<50} {'Pass Rate':>10} {'Passed':>8}")
    print("-" * 70)
    for result in sorted_results:
        print(f"{result['model_name']:<50} {result['pass_rate']:>9.1f}% {result['passed_tests']:>3}/{result['total_tests']:<3}")

    print("\n" + "=" * 70)
    print("✅ Comparison report generated successfully!")
    print("=" * 70)


if __name__ == "__main__":
    main()

